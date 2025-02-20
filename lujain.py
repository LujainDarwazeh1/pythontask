import openpyxl as xl
import xml.etree.ElementTree as ET


wb = xl.load_workbook('taskexcel_1.xlsx')
sheet = wb['Sheet1']


def convert_to_dollar():
    price_dict = {}
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        if isinstance(cell.value, (int, float)):
            sku_cell = sheet.cell(row, 1)
            sku_code = sku_cell.value
            convert_value = round(cell.value * 1.28, 2)
            convert_cell = sheet.cell(row, 4)
            convert_cell.value = convert_value
            price_dict[sku_code] = convert_value
    wb.save('taskexcel_2.xlsx')
    return price_dict


def update_xml(price_dict):

    tree = ET.parse('task.xml')
    root = tree.getroot()


    namespace = {'ns': 'http://www.demandware.com/xml/impex/pricebook/2006-10-31'
                 }


    for price_table in root.findall('.//ns:price-table', namespace):
        product_id = price_table.get('product-id')

        if product_id in price_dict:

            amount = price_table.find('ns:amount', namespace)
            amount.text = str(price_dict[product_id])


    tree.write('update.xml')


price_dict = convert_to_dollar()
update_xml(price_dict)
